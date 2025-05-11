from openpyxl import Workbook, load_workbook
import os

# File path
file_path = 'HW_list.xlsx'

# Check if file exists, else create a new workbook
if os.path.exists(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    # Get the last serial number used from the last row
    last_serial_number = ws.max_row
else:
    wb = Workbook()
    ws = wb.active
    # Add headers
    ws.append(["S.No", "Model Name", "Series"])
    last_serial_number = 1  # Start with 1 if new file

print(f"✅ Excel file is ready. Continuing from serial number {last_serial_number}.\n")

# Keep taking model names and series until user types 'exit'
while True:
    model_name = input(f"Enter Model Name {last_serial_number}: ")
    if model_name.lower() == 'exit':
        break
    series = input("Enter Series: ")
    if series.lower() == 'm':
        series = 'Mainlines'
    if series.lower() == 'uh':
        series = 'Ultra Hots'
    if series.lower() == 'ssbmw':
        series = 'Silver Series BMW'
    if series.lower() == 'ppc':
        series = 'Premiums Pop Culture'
    if series.lower() == 'ns':
        series = 'Neon Speeders'
    if series.lower() == 'ssff':
        series = 'Silver Series Fast & Furious'
    ws.append([last_serial_number, model_name, series])
    
    # Increment serial number
    last_serial_number += 1

# Save the file
wb.save(file_path)

print("\n✅ All model names and series have been saved to the Excel file!")

