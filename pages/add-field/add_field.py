#!/usr/bin/env python3
"""
DieCastTracker - Add Field Page Utilities
Utility functions for adding new fields/columns to the collection
"""

import os
import sys

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from utils.backup_utils import create_backup
from openpyxl import load_workbook

# Path to the Excel file
EXCEL_FILE_PATH = os.path.join("data", "HW_list.xlsx")

def add_field(field_name: str):
    """Add a new field/column to the Excel file"""
    try:
        # Create backup before adding field
        if not create_backup(EXCEL_FILE_PATH):
            raise Exception("Failed to create backup")
        
        # Validate field name
        field_name = field_name.strip()
        if not field_name:
            raise Exception("Field name cannot be empty")
        
        if len(field_name) > 50:
            raise Exception("Field name is too long (max 50 characters)")
        
        # Check for invalid characters
        invalid_chars = ['/', '\\', '?', '*', '[', ']', ':', ';']
        if any(char in field_name for char in invalid_chars):
            raise Exception(f"Field name contains invalid characters: {', '.join(invalid_chars)}")
        
        # Load workbook
        if not os.path.exists(EXCEL_FILE_PATH):
            raise Exception("Excel file not found")
        
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb.active
        
        # Check if field already exists
        headers = [cell.value for cell in ws[1]]
        if field_name in headers:
            raise Exception(f"Field '{field_name}' already exists")
        
        # Add the new field
        new_column = len(headers) + 1
        ws.cell(row=1, column=new_column, value=field_name)
        
        # Save workbook
        wb.save(EXCEL_FILE_PATH)
        
        return {
            "success": True,
            "column_index": new_column,
            "message": f"Field '{field_name}' added successfully!"
        }
    except Exception as e:
        raise Exception(f"Error adding field: {str(e)}")
