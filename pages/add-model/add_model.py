#!/usr/bin/env python3
"""
DieCastTracker - Add Model Page Utilities
Utility functions for adding models to the collection
"""

import os
import sys

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from utils.backup_utils import create_backup
from openpyxl import load_workbook

# Path to the Excel file
EXCEL_FILE_PATH = os.path.join("data", "HW_list.xlsx")

def add_model(model_name: str, series: str, subseries: str):
    """Add a new model to the Excel file"""
    try:
        # Create backup before adding (if file exists)
        if os.path.exists(EXCEL_FILE_PATH):
            if not create_backup(EXCEL_FILE_PATH):
                raise Exception("Failed to create backup")
        
        # Load existing workbook or create new one
        if os.path.exists(EXCEL_FILE_PATH):
            wb = load_workbook(EXCEL_FILE_PATH)
            ws = wb.active
            last_serial_number = ws.max_row
        else:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(["S.No", "Model Name", "Series"])
            last_serial_number = 1
        
        # Add new row (using subseries as the series field in Excel)
        ws.append([
            last_serial_number,
            model_name.strip(),
            subseries
        ])
        
        # Save workbook
        wb.save(EXCEL_FILE_PATH)
        
        return {
            "success": True,
            "serial_number": last_serial_number,
            "message": f"Successfully added '{model_name}' to the collection!"
        }
    except Exception as e:
        raise Exception(f"Error adding model: {str(e)}")
