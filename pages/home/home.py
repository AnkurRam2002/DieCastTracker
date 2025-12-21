#!/usr/bin/env python3
"""
DieCastTracker - Home Page Utilities
Utility functions for home page operations (view, search, update, delete)
"""

import os
import sys

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from utils.backup_utils import create_backup
from openpyxl import load_workbook

# Path to the Excel file
EXCEL_FILE_PATH = os.path.join("data", "HW_list.xlsx")

def load_excel_data():
    """Load data from the Excel file"""
    try:
        if os.path.exists(EXCEL_FILE_PATH):
            import pandas as pd
            df = pd.read_excel(EXCEL_FILE_PATH)
            df = df.fillna("")
            return df
        else:
            return None
    except Exception as e:
        raise Exception(f"Error loading Excel file: {str(e)}")

def update_model(serial_number: int, updates: dict):
    """Update a model in the Excel file"""
    try:
        # Create backup before update
        if not create_backup(EXCEL_FILE_PATH):
            raise Exception("Failed to create backup")
        
        # Load existing workbook
        if not os.path.exists(EXCEL_FILE_PATH):
            raise Exception("Excel file not found")
        
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb.active
        
        # Find the row with the given serial number
        target_row = None
        for row_num in range(2, ws.max_row + 1):
            if ws.cell(row=row_num, column=1).value == serial_number:
                target_row = row_num
                break
        
        if target_row is None:
            raise Exception(f"Model with serial number {serial_number} not found")
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Update the fields that are provided
        for field_name, new_value in updates.items():
            if field_name in headers:
                col_index = headers.index(field_name) + 1
                ws.cell(row=target_row, column=col_index, value=str(new_value).strip() if new_value else "")
        
        # Save workbook
        wb.save(EXCEL_FILE_PATH)
        return True
    except Exception as e:
        raise Exception(f"Error updating model: {str(e)}")

def delete_model(serial_number: int):
    """Delete a model from the Excel file"""
    try:
        # Create backup before deletion
        if not create_backup(EXCEL_FILE_PATH):
            raise Exception("Failed to create backup")
        
        # Load existing workbook
        if not os.path.exists(EXCEL_FILE_PATH):
            raise Exception("Excel file not found")
        
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb.active
        
        # Find the row with the given serial number
        target_row = None
        for row_num in range(2, ws.max_row + 1):
            if ws.cell(row=row_num, column=1).value == serial_number:
                target_row = row_num
                break
        
        if target_row is None:
            raise Exception(f"Model with serial number {serial_number} not found")
        
        # Delete the row
        ws.delete_rows(target_row)
        
        # Renumber serial numbers
        for row_num in range(2, ws.max_row + 1):
            ws.cell(row=row_num, column=1, value=row_num - 1)
        
        # Save workbook
        wb.save(EXCEL_FILE_PATH)
        return True
    except Exception as e:
        raise Exception(f"Error deleting model: {str(e)}")

def search_models(query: str):
    """Search through the data"""
    try:
        import pandas as pd
        df = load_excel_data()
        if df is None:
            return []
        
        if not query:
            # Return all data if no search query
            return df.to_dict('records')
        else:
            # Search across all text columns
            mask = pd.Series([False] * len(df))
            for col in df.select_dtypes(include=['object']).columns:
                mask |= df[col].astype(str).str.contains(query, case=False, na=False)
            
            filtered_df = df[mask]
            return filtered_df.to_dict('records')
    except Exception as e:
        raise Exception(f"Error searching models: {str(e)}")

