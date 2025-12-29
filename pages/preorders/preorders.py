#!/usr/bin/env python3
"""
DieCastTracker - Preorders Page Utilities
Utility functions for managing preorders
"""

import os
import sys
from datetime import datetime
import pandas as pd
import math

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from utils.backup_utils import create_backup
from openpyxl import load_workbook, Workbook

# Path to the preorders Excel file
PREORDERS_FILE_PATH = os.path.join("data", "preorders.xlsx")

def load_preorders_data():
    """Load data from the preorders Excel file"""
    try:
        if os.path.exists(PREORDERS_FILE_PATH):
            df = pd.read_excel(PREORDERS_FILE_PATH)
            # Replace NaN, inf, and -inf values with empty strings or 0 for JSON compatibility
            df = df.replace([float('inf'), float('-inf')], '')
            df = df.fillna("")
            
            # Clean all columns to ensure JSON compatibility
            for col in df.columns:
                # Replace inf/-inf/NaN with empty string for all columns
                df[col] = df[col].replace([float('inf'), float('-inf')], '')
                df[col] = df[col].fillna('')
                
                # For numeric columns, handle conversion carefully
                if df[col].dtype in ['float64', 'int64', 'float32', 'int32']:
                    # Convert to string, handling any remaining issues
                    df[col] = df[col].apply(lambda x: '' if pd.isna(x) or (isinstance(x, float) and (math.isinf(x) or math.isnan(x))) else str(x) if x != '' else '')
                else:
                    # For text columns, ensure they're strings and clean
                    df[col] = df[col].astype(str).replace('nan', '').replace('None', '')
            
            return df
        else:
            return None
    except Exception as e:
        raise Exception(f"Error loading preorders file: {str(e)}")

def add_preorder(seller, models, eta, total_price, po_amount, on_arrival_amount, delivery_status=None):
    """Add a new preorder to the Excel file"""
    try:
        # Create backup before adding (if file exists)
        if os.path.exists(PREORDERS_FILE_PATH):
            if not create_backup(PREORDERS_FILE_PATH):
                raise Exception("Failed to create backup")
        
        # Load existing workbook or create new one
        if os.path.exists(PREORDERS_FILE_PATH):
            wb = load_workbook(PREORDERS_FILE_PATH)
            ws = wb.active
            # Check if we need to migrate from old "Status" column to "Delivery Status"
            headers = [cell.value for cell in ws[1]]
            if "Status" in headers and "Delivery Status" not in headers:
                # Migrate old column name to new one
                status_col = headers.index("Status") + 1
                ws.cell(row=1, column=status_col, value="Delivery Status")
                wb.save(PREORDERS_FILE_PATH)
        else:
            wb = Workbook()
            ws = wb.active
            # Add headers
            ws.append(["S.No", "Seller", "Models", "ETA", "Total Price", "PO Amount", "On Arrival Amount", "Delivery Status", "Date Added"])
        
        # Get next serial number
        if ws.max_row == 1:
            serial_number = 1
        else:
            serial_number = ws.max_row
        
        # Set default delivery status if not provided
        if not delivery_status:
            delivery_status = "Pending"
        
        # Format ETA as month (YYYY-MM) if provided
        eta_formatted = ""
        if eta:
            # If it's already in YYYY-MM format, use it; otherwise try to parse
            if len(str(eta)) == 7 and str(eta).count('-') == 1:
                eta_formatted = str(eta)
            else:
                try:
                    # Try to parse as date and extract month
                    eta_date = datetime.strptime(str(eta), "%Y-%m-%d")
                    eta_formatted = eta_date.strftime("%Y-%m")
                except:
                    eta_formatted = str(eta)
        
        # Add new row
        ws.append([
            serial_number,
            seller.strip() if seller else "",
            models.strip() if models else "",
            eta_formatted,
            total_price if total_price else "",
            po_amount if po_amount else "",
            on_arrival_amount if on_arrival_amount else "",
            delivery_status,
            datetime.now().strftime("%Y-%m-%d")
        ])
        
        # Save workbook
        wb.save(PREORDERS_FILE_PATH)
        
        return {
            "success": True,
            "serial_number": serial_number,
            "message": f"Successfully added preorder #{serial_number}!"
        }
    except Exception as e:
        raise Exception(f"Error adding preorder: {str(e)}")

def update_preorder(serial_number, updates):
    """Update a preorder in the Excel file"""
    try:
        # Create backup before update
        if not create_backup(PREORDERS_FILE_PATH):
            raise Exception("Failed to create backup")
        
        # Load existing workbook
        if not os.path.exists(PREORDERS_FILE_PATH):
            raise Exception("Preorders file not found")
        
        wb = load_workbook(PREORDERS_FILE_PATH)
        ws = wb.active
        
        # Find the row with the given serial number
        target_row = None
        for row_num in range(2, ws.max_row + 1):
            if ws.cell(row=row_num, column=1).value == serial_number:
                target_row = row_num
                break
        
        if target_row is None:
            raise Exception(f"Preorder with serial number {serial_number} not found")
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Migrate old "Status" column to "Delivery Status" if needed
        if "Status" in headers and "Delivery Status" not in headers:
            status_col = headers.index("Status") + 1
            ws.cell(row=1, column=status_col, value="Delivery Status")
            headers[status_col - 1] = "Delivery Status"
            wb.save(PREORDERS_FILE_PATH)
        
        # Update the fields that are provided
        for field_name, new_value in updates.items():
            # Handle both old and new column names
            if field_name in headers:
                col_index = headers.index(field_name) + 1
                # Format ETA as month if it's the ETA field
                if field_name == "ETA" and new_value:
                    # Format as YYYY-MM if it's a date or already a month
                    if len(str(new_value)) == 7 and str(new_value).count('-') == 1:
                        formatted_value = str(new_value)
                    else:
                        try:
                            # Try to parse as date and extract month
                            eta_date = datetime.strptime(str(new_value), "%Y-%m-%d")
                            formatted_value = eta_date.strftime("%Y-%m")
                        except:
                            formatted_value = str(new_value).strip()
                    ws.cell(row=target_row, column=col_index, value=formatted_value)
                else:
                    ws.cell(row=target_row, column=col_index, value=str(new_value).strip() if new_value else "")
            elif field_name == "Delivery Status" and "Status" in headers:
                # If updating Delivery Status but file has old "Status" column, migrate it
                status_col = headers.index("Status") + 1
                ws.cell(row=1, column=status_col, value="Delivery Status")
                ws.cell(row=target_row, column=status_col, value=str(new_value).strip() if new_value else "")
                headers[status_col - 1] = "Delivery Status"
        
        # Update delivery status if provided (after migration check)
        if "Delivery Status" in updates and "Delivery Status" in headers:
            delivery_status_col = headers.index("Delivery Status") + 1
            ws.cell(row=target_row, column=delivery_status_col, value=updates.get("Delivery Status"))
        
        # Save workbook
        wb.save(PREORDERS_FILE_PATH)
        return True
    except Exception as e:
        raise Exception(f"Error updating preorder: {str(e)}")

def delete_preorder(serial_number):
    """Delete a preorder from the Excel file"""
    try:
        # Create backup before deletion
        if not create_backup(PREORDERS_FILE_PATH):
            raise Exception("Failed to create backup")
        
        # Load existing workbook
        if not os.path.exists(PREORDERS_FILE_PATH):
            raise Exception("Preorders file not found")
        
        wb = load_workbook(PREORDERS_FILE_PATH)
        ws = wb.active
        
        # Find the row with the given serial number
        target_row = None
        for row_num in range(2, ws.max_row + 1):
            if ws.cell(row=row_num, column=1).value == serial_number:
                target_row = row_num
                break
        
        if target_row is None:
            raise Exception(f"Preorder with serial number {serial_number} not found")
        
        # Delete the row
        ws.delete_rows(target_row)
        
        # Renumber serial numbers
        for row_num in range(2, ws.max_row + 1):
            ws.cell(row=row_num, column=1, value=row_num - 1)
        
        # Save workbook
        wb.save(PREORDERS_FILE_PATH)
        return True
    except Exception as e:
        raise Exception(f"Error deleting preorder: {str(e)}")

def get_preorders_statistics():
    """Get statistics about preorders"""
    try:
        df = load_preorders_data()
        if df is None or df.empty:
            return {
                "total_preorders": 0,
                "total_value": 0,
                "total_po_amount": 0,
                "total_on_arrival": 0,
                "payment_done": 0,
                "payment_remaining": 0,
                "status_breakdown": {},
                "upcoming_arrivals": []
            }
        
        # Basic statistics
        total_preorders = len(df)
        
        # Calculate totals
        total_value = 0
        total_po_amount = 0
        total_on_arrival = 0
        
        # Helper function to safely convert to float
        def safe_float(value):
            if pd.isna(value) or value == '' or value is None:
                return 0.0
            try:
                # If it's already a number, return it
                if isinstance(value, (int, float)):
                    return float(value)
                # Convert to string and clean
                value_str = str(value).strip()
                # Remove any currency symbols, commas, and whitespace
                value_str = value_str.replace('₹', '').replace(',', '').replace(' ', '').replace('$', '').strip()
                if value_str == '' or value_str == '-' or value_str.lower() == 'nan':
                    return 0.0
                # Try to convert to float
                return float(value_str)
            except (ValueError, TypeError) as e:
                return 0.0
        
        # Calculate totals - handle column names with potential whitespace
        # Create a mapping of normalized column names
        column_map = {col.strip(): col for col in df.columns}
        
        # Find and calculate Total Price
        if "Total Price" in column_map:
            total_value = df[column_map["Total Price"]].apply(safe_float).sum()
        elif "TotalPrice" in column_map:
            total_value = df[column_map["TotalPrice"]].apply(safe_float).sum()
        else:
            # Try to find column by partial match
            for col in df.columns:
                if "total" in str(col).lower() and "price" in str(col).lower():
                    total_value = df[col].apply(safe_float).sum()
                    break
            
        # Find and calculate PO Amount
        if "PO Amount" in column_map:
            total_po_amount = df[column_map["PO Amount"]].apply(safe_float).sum()
        elif "POAmount" in column_map:
            total_po_amount = df[column_map["POAmount"]].apply(safe_float).sum()
        else:
            # Try to find column by partial match
            for col in df.columns:
                if "po" in str(col).lower() and "amount" in str(col).lower():
                    total_po_amount = df[col].apply(safe_float).sum()
                    break
                    
        # Find and calculate On Arrival Amount
        on_arrival_col = None
        if "On Arrival Amount" in column_map:
            on_arrival_col = column_map["On Arrival Amount"]
            total_on_arrival = df[on_arrival_col].apply(safe_float).sum()
        elif "OnArrivalAmount" in column_map:
            on_arrival_col = column_map["OnArrivalAmount"]
            total_on_arrival = df[on_arrival_col].apply(safe_float).sum()
        else:
            # Try to find column by partial match
            for col in df.columns:
                if "arrival" in str(col).lower() and "amount" in str(col).lower():
                    on_arrival_col = col
                    total_on_arrival = df[col].apply(safe_float).sum()
                    break
        
        # Calculate Payment Done and Payment Remaining
        # Payment Done = All PO Amount + On Arrival Amount for Shipped/Delivered items
        # Payment Remaining = On Arrival Amount for Pending items
        payment_done = total_po_amount  # Start with all PO amounts
        payment_remaining = 0
        
        # Handle both old and new column names for delivery status
        status_col = "Delivery Status" if "Delivery Status" in df.columns else "Status"
        
        if on_arrival_col:
            for idx, row in df.iterrows():
                on_arrival = safe_float(row.get(on_arrival_col, 0))
                delivery_status = str(row.get(status_col, "Pending")).strip()
                
                # If status is Shipped or Delivered, add on_arrival to payment_done
                if delivery_status.lower() in ["shipped", "delivered"]:
                    payment_done += on_arrival
                # If status is Pending, add to payment_remaining
                elif delivery_status.lower() == "pending":
                    payment_remaining += on_arrival
        
        # Delivery status breakdown (handle both old and new column names)
        status_breakdown = {}
        if "Delivery Status" in df.columns:
            status_counts = df["Delivery Status"].value_counts()
            status_breakdown = status_counts.to_dict()
        elif "Status" in df.columns:
            # Use old column name if new one doesn't exist
            status_counts = df["Status"].value_counts()
            status_breakdown = status_counts.to_dict()
        
        # Upcoming arrivals (current month and next month)
        upcoming_arrivals = []
        if "ETA" in df.columns:
            today = datetime.now()
            current_month = today.strftime("%Y-%m")
            # Calculate next month
            if today.month == 12:
                next_month = f"{today.year + 1}-01"
            else:
                next_month = f"{today.year}-{today.month + 1:02d}"
            # Handle both old and new column names
            status_col = "Delivery Status" if "Delivery Status" in df.columns else "Status"
            for idx, row in df.iterrows():
                eta = row.get("ETA")
                delivery_status = row.get(status_col, "Pending")
                # Only show pending or shipped items
                if delivery_status not in ["Delivered"]:
                    if eta:
                        try:
                            # ETA is now in YYYY-MM format
                            eta_month = str(eta)[:7]  # Get YYYY-MM part
                            if eta_month == current_month or eta_month == next_month:
                                upcoming_arrivals.append({
                                    "serial": row.get("S.No", "N/A"),
                                    "models": row.get("Models", "N/A"),
                                    "eta": eta_month,
                                    "month": eta_month,
                                    "seller": row.get("Seller", "N/A"),
                                    "status": delivery_status
                                })
                        except:
                            pass
        
        return {
            "total_preorders": total_preorders,
            "total_value": round(total_value, 2),
            "total_po_amount": round(total_po_amount, 2),
            "total_on_arrival": round(total_on_arrival, 2),
            "payment_done": round(payment_done, 2),
            "payment_remaining": round(payment_remaining, 2),
            "status_breakdown": status_breakdown,
            "upcoming_arrivals": sorted(upcoming_arrivals, key=lambda x: x.get("month", x.get("eta", "")))
        }
    except Exception as e:
        raise Exception(f"Error getting preorders statistics: {str(e)}")

