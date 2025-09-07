from openpyxl import load_workbook
import os

# File path
import os
file_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', 'HW_list.xlsx')

# Check if file exists
if not os.path.exists(file_path):
    print("❌ Excel file not found! Please make sure 'data/HW_list.xlsx' exists.")
    exit()

# Load workbook and worksheet
wb = load_workbook(file_path)
ws = wb.active

# Show current headers
headers = [cell.value for cell in ws[1]]
print("📑 Current fields (columns):")
for i, header in enumerate(headers, start=1):
    print(f"{i}. {header}")

# Ask for new field name
new_field = input("\nEnter the name of the new field you want to add: ")

# Check if the field already exists
if new_field in headers:
    print(f"⚠️ The field '{new_field}' already exists. No changes made.")
else:
    ws.cell(row=1, column=len(headers) + 1, value=new_field)
    print(f"✅ Field '{new_field}' added as column {len(headers) + 1}.")

    # Save the file
    wb.save(file_path)
    print("✅ Excel file updated successfully.")

