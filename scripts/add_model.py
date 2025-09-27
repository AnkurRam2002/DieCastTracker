from openpyxl import Workbook, load_workbook
import os

# File path
import os
file_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', 'HW_list.xlsx')

# Check if file exists, else create a new workbook
if os.path.exists(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    # Get the last serial number used from the last row
    last_serial_number = ws.max_row
else:
    wb = Workbook()
    ws = wb.active
    # Add headers
    ws.append(["S.No", "Model Name", "Brand", "Series", "Subseries"])
    last_serial_number = 1  # Start with 1 if new file

print(f"✅ Excel file is ready. Continuing from serial number {last_serial_number}.\n")

# Brand options
brand_options = ["Hot Wheels", "Matchbox", "Other"]

# Series options as a nested dictionary
series_options = {
    "Mainlines": [
        "Mainlines",
        "57th Anniversary Series"
    ],
    "Semi-Premiums": [
        "Ultra Hots",
        "Silver Series BMW",
        "Silver Series Fast & Furious Villains",
        "HW Speed Graphics",
        "Neon Speeders",
        "1/4 Mile Finals Series",
        "Fast & Furious Hobbs & Shaw"
    ],
    "Premiums": [
        "Premiums Pop Culture",
        "Premiums Boulevard"
    ]
}

# Function to display brand menu
def display_brand_menu():
    print("\nSelect Brand:")
    print("-" * 20)
    for i, brand in enumerate(brand_options, start=1):
        print(f"{i}. {brand}")
    print("-" * 20)

# Function to display the series menu
def display_series_menu():
    print("\nSelect Main Series and Sub Series:")
    print("-" * 40)
    for i, (main_series, sub_series_list) in enumerate(series_options.items(), start=1):
        print(f"{i}. {main_series}")
        for j, sub_series in enumerate(sub_series_list, start=1):
            print(f"   {j}. {sub_series}")
        print()
    print("-" * 40)

# Keep taking model names and series until user types 'exit'
while True:
    model_input = input(f"\nEnter Model Name {last_serial_number} (or type 'exit' to finish): ")
    if model_input.lower() == 'exit':
        break

    # Select brand first
    display_brand_menu()
    while True:
        try:
            brand_choice = int(input("Enter Brand number: "))
            if 1 <= brand_choice <= len(brand_options):
                selected_brand = brand_options[brand_choice - 1]
                break
            else:
                print("❌ Invalid Brand number.")
        except ValueError:
            print("❌ Please enter a valid number.")

    # Check for shortcut e.g. 'Car Name#13'
    if '#' in model_input:
        model_name, shortcut = model_input.split('#', 1)
        if len(shortcut) == 2 and shortcut.isdigit():
            main_choice = int(shortcut[0])
            sub_choice = int(shortcut[1])
            try:
                main_series_name = list(series_options.keys())[main_choice - 1]
                sub_series_list = series_options[main_series_name]
                selected_subseries = sub_series_list[sub_choice - 1]
                selected_series = main_series_name
                print(f"✅ Auto-selected Series: {selected_series} - {selected_subseries}")
            except (IndexError, ValueError):
                print("❌ Invalid shortcut code. Falling back to manual selection.")
                display_series_menu()
                main_choice = int(input("Enter Main Series number: "))
                main_series_name = list(series_options.keys())[main_choice - 1]
                sub_series_list = series_options[main_series_name]
                sub_choice = int(input("Enter Sub Series number: "))
                selected_series = main_series_name
                selected_subseries = sub_series_list[sub_choice - 1]
        else:
            print("❌ Invalid shortcut format. Use two digits like #12.")
            continue
    else:
        model_name = model_input
        display_series_menu()

        while True:
            try:
                main_choice = int(input("Enter Main Series number: "))
                if 1 <= main_choice <= len(series_options):
                    main_series_name = list(series_options.keys())[main_choice - 1]
                    selected_series = main_series_name
                    break
                else:
                    print("❌ Invalid Main Series number.")
            except ValueError:
                print("❌ Please enter a valid number.")

        sub_series_list = series_options[main_series_name]
        while True:
            try:
                sub_choice = int(input("Enter Sub Series number: "))
                if 1 <= sub_choice <= len(sub_series_list):
                    selected_subseries = sub_series_list[sub_choice - 1]
                    break
                else:
                    print("❌ Invalid Sub Series number.")
            except ValueError:
                print("❌ Please enter a valid number.")

    # Append to sheet: S.No, Model Name, Brand, Series, Subseries
    ws.append([last_serial_number, model_name.strip(), selected_brand, selected_series, selected_subseries])

    # Increment serial number
    last_serial_number += 1

# Save the file
wb.save(file_path)

print("\n✅ All model names and series have been saved to the Excel file!")



