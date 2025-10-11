#!/usr/bin/env python3
"""
DieCastTracker - CLI Utilities
Shared utilities for CLI scripts to reduce code duplication
"""

import os
import sys
from openpyxl import Workbook, load_workbook
from typing import List, Tuple, Optional, Dict, Any
from datetime import datetime

# File path configuration
FILE_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', 'HW_list.xlsx')

# Import series configuration from centralized file
from series_config import SERIES_OPTIONS

class ExcelManager:
    """Manages Excel file operations"""
    
    @staticmethod
    def load_workbook_data() -> Tuple[Optional[Any], Optional[Any], List[str], List[List]]:
        """Load workbook and return workbook, worksheet, headers, and data"""
        if not os.path.exists(FILE_PATH):
            print("❌ Excel file not found! Please make sure 'data/HW_list.xlsx' exists.")
            return None, None, [], []
        
        try:
            wb = load_workbook(FILE_PATH)
            ws = wb.active
            
            # Get headers
            headers = [cell.value for cell in ws[1]]
            
            # Get all data rows
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    data.append(list(row))
            
            return wb, ws, headers, data
        except Exception as e:
            print(f"❌ Error loading Excel file: {e}")
            return None, None, [], []
    
    @staticmethod
    def create_new_workbook() -> Tuple[Any, Any]:
        """Create a new workbook with default headers"""
        wb = Workbook()
        ws = wb.active
        ws.append(["S.No", "Model Name", "Series"])
        return wb, ws
    
    @staticmethod
    def save_workbook(wb: Any) -> bool:
        """Save workbook to file"""
        try:
            wb.save(FILE_PATH)
            return True
        except Exception as e:
            print(f"❌ Error saving Excel file: {e}")
            return False

class DisplayUtils:
    """Utility functions for displaying data"""
    
    @staticmethod
    def print_header(title: str, width: int = 50) -> None:
        """Print a formatted header"""
        print(f"\n{title}")
        print("=" * width)
    
    @staticmethod
    def print_section(title: str, width: int = 30) -> None:
        """Print a formatted section header"""
        print(f"\n{title}")
        print("-" * width)
    
    @staticmethod
    def display_models_table(data: List[List], headers: List[str], title: str = "Collection") -> None:
        """Display models in a formatted table"""
        if not data:
            print(f"📭 No models found in {title.lower()}!")
            return
        
        print(f"\n📋 {title} ({len(data)} models):")
        print("=" * 80)
        
        # Create header row
        header_row = f"{'S.No':<6} {'Model Name':<30} {'Series':<25}"
        if len(headers) > 3:
            header_row += f" {'Brand':<15}"
        print(header_row)
        print("-" * 80)
        
        # Display data rows
        for row in data:
            serial_no = row[0] if len(row) > 0 else "N/A"
            model_name = row[1] if len(row) > 1 else "N/A"
            series = row[2] if len(row) > 2 else "N/A"
            
            data_row = f"{serial_no:<6} {model_name:<30} {series:<25}"
            if len(row) > 3:
                brand = row[3] if len(row) > 3 else "N/A"
                data_row += f" {brand:<15}"
            print(data_row)
    
    @staticmethod
    def display_series_menu() -> None:
        """Display the series selection menu"""
        print("\nSelect Main Series and Sub Series:")
        print("-" * 40)
        for i, (main_series, sub_series_list) in enumerate(SERIES_OPTIONS.items(), start=1):
            print(f"{i}. {main_series}")
            for j, sub_series in enumerate(sub_series_list, start=1):
                print(f"   {j}. {sub_series}")
            print()
        print("-" * 40)

class InputUtils:
    """Utility functions for user input"""
    
    @staticmethod
    def get_choice(prompt: str, valid_choices: List[str]) -> str:
        """Get user choice with validation"""
        while True:
            choice = input(prompt).strip()
            if choice in valid_choices:
                return choice
            print(f"❌ Invalid choice! Please enter one of: {', '.join(valid_choices)}")
    
    @staticmethod
    def get_serial_number(prompt: str = "Enter Serial Number: ") -> Optional[int]:
        """Get serial number with validation"""
        while True:
            try:
                serial_no = int(input(prompt).strip())
                if serial_no > 0:
                    return serial_no
                print("❌ Serial number must be positive!")
            except ValueError:
                print("❌ Please enter a valid number!")
            except KeyboardInterrupt:
                return None
    
    @staticmethod
    def confirm_action(message: str, default: bool = False) -> bool:
        """Get user confirmation for an action"""
        default_text = "Y/n" if default else "y/N"
        while True:
            response = input(f"{message} ({default_text}): ").strip().lower()
            if not response:
                return default
            if response in ['y', 'yes']:
                return True
            elif response in ['n', 'no']:
                return False
            print("❌ Please enter 'y' or 'n'")

class SearchUtils:
    """Utility functions for searching"""
    
    @staticmethod
    def search_models(data: List[List], search_term: str) -> List[Tuple[int, List]]:
        """Search for models matching the search term"""
        matches = []
        search_lower = search_term.lower()
        
        for i, row in enumerate(data):
            if any(search_lower in str(cell).lower() for cell in row):
                matches.append((i, row))
        
        return matches
    
    @staticmethod
    def find_model_by_serial(data: List[List], serial_no: int) -> Tuple[Optional[int], Optional[List]]:
        """Find model by serial number"""
        for i, row in enumerate(data):
            if row[0] == serial_no:
                return i, row
        return None, None

class ModelUtils:
    """Utility functions for model operations"""
    
    @staticmethod
    def parse_shortcut_input(model_input: str) -> Tuple[str, Optional[str], Optional[str]]:
        """Parse shortcut input like 'Car Name#13'"""
        if '#' not in model_input:
            return model_input, None, None
        
        try:
            model_name, shortcut = model_input.split('#', 1)
            if len(shortcut) == 2 and shortcut.isdigit():
                main_choice = int(shortcut[0])
                sub_choice = int(shortcut[1])
                
                main_series_list = list(SERIES_OPTIONS.keys())
                if 1 <= main_choice <= len(main_series_list):
                    main_series_name = main_series_list[main_choice - 1]
                    sub_series_list = SERIES_OPTIONS[main_series_name]
                    if 1 <= sub_choice <= len(sub_series_list):
                        selected_subseries = sub_series_list[sub_choice - 1]
                        return model_name.strip(), main_series_name, selected_subseries
            
            print("❌ Invalid shortcut format. Use two digits like #12.")
            return model_name, None, None
        except (ValueError, IndexError):
            print("❌ Invalid shortcut code. Falling back to manual selection.")
            return model_input, None, None
    
    @staticmethod
    def get_series_selection() -> Tuple[str, str]:
        """Get series selection from user"""
        DisplayUtils.display_series_menu()
        
        # Get main series
        main_series_list = list(SERIES_OPTIONS.keys())
        while True:
            try:
                main_choice = int(input("Enter Main Series number: "))
                if 1 <= main_choice <= len(main_series_list):
                    main_series_name = main_series_list[main_choice - 1]
                    break
                print("❌ Invalid Main Series number.")
            except ValueError:
                print("❌ Please enter a valid number.")
        
        # Get sub series
        sub_series_list = SERIES_OPTIONS[main_series_name]
        while True:
            try:
                sub_choice = int(input("Enter Sub Series number: "))
                if 1 <= sub_choice <= len(sub_series_list):
                    selected_subseries = sub_series_list[sub_choice - 1]
                    break
                print("❌ Invalid Sub Series number.")
            except ValueError:
                print("❌ Please enter a valid number.")
        
        return main_series_name, selected_subseries

class BackupUtils:
    """Utility functions for backup operations"""
    
    @staticmethod
    def create_backup_if_needed() -> bool:
        """Create backup if backup_utils is available"""
        try:
            from backup_utils import create_backup
            return create_backup(FILE_PATH)
        except ImportError:
            print("⚠️  Backup utilities not available. Proceeding without backup.")
            return True
        except Exception as e:
            print(f"⚠️  Backup failed: {e}. Proceeding without backup.")
            return True

def print_success(message: str) -> None:
    """Print success message"""
    print(f"✅ {message}")

def print_error(message: str) -> None:
    """Print error message"""
    print(f"❌ {message}")

def print_warning(message: str) -> None:
    """Print warning message"""
    print(f"⚠️  {message}")

def print_info(message: str) -> None:
    """Print info message"""
    print(f"ℹ️  {message}")

def pause_for_user() -> None:
    """Pause and wait for user input"""
    input("\n⏸️  Press Enter to continue...")

def clear_screen() -> None:
    """Clear the screen"""
    os.system('cls' if os.name == 'nt' else 'clear')
