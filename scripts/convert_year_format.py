#!/usr/bin/env python3
"""
DieCastTracker - Year Format Converter
Converts year formats in model names from "12" or "1012" to "'12" format
"""

import os
import sys
import re
import pandas as pd
from openpyxl import load_workbook

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from utils.backup_utils import create_backup

# Path to the Excel file
EXCEL_FILE_PATH = os.path.join("data", "HW_list.xlsx")

def convert_year_format(text):
    """
    Convert year formats in text from "12" or "1012" to "'12" format
    
    Examples:
    - "2020 Mustang" -> "'20 Mustang"
    - "2012 Camaro" -> "'12 Camaro"
    - "12 Corvette" -> "'12 Corvette"
    - "1012 Ferrari" -> "'12 Ferrari"
    - "'12 Porsche" -> "'12 Porsche" (already correct, no change)
    """
    if not isinstance(text, str) or not text.strip():
        return text
    
    # Pattern 1: Four-digit years (1900-2099, and also 1000-1899 for edge cases like "1012")
    # Matches: "2020", "2012", "1012", "1995", etc.
    def replace_four_digit_year(match):
        year = match.group(0)
        # Extract last two digits
        last_two = year[-2:]
        # Check if it's already in 'XX format
        pos = match.start()
        if pos > 0 and text[pos - 1] == "'":
            return year  # Already formatted
        return f"'{last_two}"
    
    # Pattern 2: Two-digit years (00-99) that are likely years
    # Only match if they appear at the start of the string or after a space,
    # and are followed by a space or end of string (to avoid matching model numbers)
    def replace_two_digit_year(match):
        year = match.group(0)
        pos = match.start()
        # Check if it's already in 'XX format
        if pos > 0 and text[pos - 1] == "'":
            return year  # Already formatted
        return f"'{year}"
    
    # First, handle four-digit years (1000-2099)
    # Match years at word boundaries, not preceded by apostrophe
    text = re.sub(r"(?<!')\b(1[0-9]{3}|20[0-9]{2})\b", replace_four_digit_year, text)
    
    # Then, handle two-digit years (00-99) that appear standalone
    # Match two digits that are:
    # - At start of string or after space/punctuation (but not after apostrophe)
    # - Followed by space, punctuation, or end of string
    # - Not part of a larger number
    # - Not already preceded by apostrophe
    # Note: The regex (?<!') already checks for apostrophe, so we're safe
    text = re.sub(r"(?<!')(?<![0-9])\b([0-9]{2})\b(?![0-9])(?=\s|$|[^\w])", replace_two_digit_year, text)
    
    return text

def convert_model_names():
    """Convert year formats in all model names"""
    try:
        # Check if file exists
        if not os.path.exists(EXCEL_FILE_PATH):
            print(f"❌ Excel file not found: {EXCEL_FILE_PATH}")
            return False
        
        # Create backup before making changes
        print("📦 Creating backup...")
        if not create_backup(EXCEL_FILE_PATH):
            print("❌ Failed to create backup. Aborting.")
            return False
        
        # Load Excel file
        print("📖 Loading Excel file...")
        df = pd.read_excel(EXCEL_FILE_PATH)
        
        # Find the Model Name column
        model_column = None
        for col in df.columns:
            if 'model' in col.lower() and 'name' in col.lower():
                model_column = col
                break
        
        if model_column is None:
            # Try to find it by position (usually second column)
            if len(df.columns) > 1:
                model_column = df.columns[1]
            else:
                print("❌ Could not find Model Name column")
                return False
        
        print(f"✅ Found Model Name column: {model_column}")
        
        # Track changes
        changes = []
        original_names = df[model_column].copy()
        
        # Convert year formats
        print("🔄 Converting year formats...")
        for idx, name in enumerate(df[model_column]):
            if pd.notna(name):
                original = str(name)
                converted = convert_year_format(original)
                if original != converted:
                    df.at[idx, model_column] = converted
                    changes.append({
                        'index': idx + 1,  # Excel row number (1-indexed, accounting for header)
                        'original': original,
                        'converted': converted
                    })
        
        # Save changes if any were made
        if changes:
            print(f"\n📝 Found {len(changes)} model names to update:")
            for change in changes[:10]:  # Show first 10
                print(f"  Row {change['index']}: '{change['original']}' -> '{change['converted']}'")
            if len(changes) > 10:
                print(f"  ... and {len(changes) - 10} more")
            
            # Save using openpyxl to preserve formatting
            print("\n💾 Saving changes...")
            wb = load_workbook(EXCEL_FILE_PATH)
            ws = wb.active
            
            # Find the column index for Model Name
            header_row = 1
            model_col_idx = None
            for col_idx, cell in enumerate(ws[header_row], start=1):
                if cell.value and 'model' in str(cell.value).lower() and 'name' in str(cell.value).lower():
                    model_col_idx = col_idx
                    break
            
            if model_col_idx is None:
                # Fallback: use second column
                model_col_idx = 2
            
            # Update the cells
            for change in changes:
                row_num = change['index'] + 1  # +1 for header row
                ws.cell(row=row_num, column=model_col_idx, value=change['converted'])
            
            wb.save(EXCEL_FILE_PATH)
            print(f"✅ Successfully updated {len(changes)} model names!")
            print(f"✅ Changes saved to {EXCEL_FILE_PATH}")
        else:
            print("ℹ️  No changes needed. All year formats are already in 'XX format.")
        
        return True
        
    except Exception as e:
        print(f"❌ Error converting year formats: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("=" * 60)
    print("🔄 Year Format Converter")
    print("=" * 60)
    print(f"Target file: {EXCEL_FILE_PATH}")
    print()
    
    success = convert_model_names()
    
    if success:
        print("\n✅ Conversion completed successfully!")
    else:
        print("\n❌ Conversion failed!")
        sys.exit(1)

