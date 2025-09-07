from openpyxl import load_workbook
import os

# File path
import os
file_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', 'HW_list.xlsx')

# Check if file exists
if not os.path.exists(file_path):
    print("❌ Excel file not found! Please make sure 'data/HW_list.xlsx' exists.")
    exit()

# Load workbook and worksheet
wb = load_workbook(file_path)
ws = wb.active

# Function to search for matching entries
def search_models(search_term):
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        serial_no, model_name, series = row
        if search_term.lower() in model_name.lower():
            results.append(row)
    return results

print("🔍 Hot Wheels Model Search 🔍")

while True:
    search_term = input("\nEnter car company or model name to search (or type 'exit' to quit): ")
    if search_term.lower() == 'exit':
        break

    matches = search_models(search_term)

    if matches:
        print(f"\n✅ Found {len(matches)} matching result(s):")
        print("-" * 50)
        for match in matches:
            print(f"S.No: {match[0]} | Model Name: {match[1]} | Series: {match[2]}")
        print("-" * 50)
    else:
        print("❌ No matching entries found.")

print("\n🔍 Search session ended. Thank you!")
