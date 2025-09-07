from openpyxl import load_workbook
import os
from collections import Counter
from datetime import datetime

# File path
import os
file_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', 'HW_list.xlsx')

def load_data():
    """Load data from Excel file"""
    if not os.path.exists(file_path):
        print("❌ Excel file not found! Please make sure 'data/HW_list.xlsx' exists.")
        return None, None
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    
    # Get all data rows
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):  # Skip empty rows
            data.append(row)
    
    return headers, data

def display_basic_stats(headers, data):
    """Display basic collection statistics"""
    print("📊 COLLECTION STATISTICS 📊")
    print("=" * 50)
    
    # Total count
    total_cars = len(data)
    print(f"🚗 Total Cars in Collection: {total_cars}")
    
    # Series breakdown
    if len(headers) >= 3:  # Assuming S.No, Model Name, Series columns
        series_column = 2  # Series is 3rd column (index 2)
        series_counts = Counter(row[series_column] for row in data if row[series_column])
        
        print(f"\n📈 Series Breakdown:")
        print("-" * 30)
        for series, count in series_counts.most_common():
            percentage = (count / total_cars) * 100 if total_cars > 0 else 0
            print(f"  {series}: {count} cars ({percentage:.1f}%)")
    
    print("=" * 50)

def display_detailed_stats(headers, data):
    """Display detailed statistics with more insights"""
    print("\n🔍 DETAILED ANALYSIS 🔍")
    print("=" * 50)
    
    total_cars = len(data)
    
    if total_cars == 0:
        print("📭 No cars in collection yet!")
        return
    
    # Series analysis
    if len(headers) >= 3:
        series_column = 2
        series_counts = Counter(row[series_column] for row in data if row[series_column])
        
        print(f"📊 Series Analysis:")
        print(f"  • Most Popular Series: {series_counts.most_common(1)[0][0]} ({series_counts.most_common(1)[0][1]} cars)")
        print(f"  • Least Popular Series: {series_counts.most_common()[-1][0]} ({series_counts.most_common()[-1][1]} cars)")
        print(f"  • Total Unique Series: {len(series_counts)}")
        
        # Series diversity
        if len(series_counts) > 1:
            diversity = len(series_counts) / total_cars * 100
            print(f"  • Series Diversity: {diversity:.1f}% (higher = more variety)")
    
    # Model name analysis
    if len(headers) >= 2:
        model_column = 1
        model_names = [row[model_column] for row in data if row[model_column]]
        
        # Find common words in model names
        all_words = []
        for name in model_names:
            if isinstance(name, str):
                all_words.extend(name.lower().split())
        
        word_counts = Counter(all_words)
        common_words = word_counts.most_common(5)
        
        print(f"\n🏷️  Model Name Insights:")
        print(f"  • Most Common Words: {', '.join([word for word, count in common_words])}")
        
        # Average model name length
        avg_length = sum(len(str(name)) for name in model_names) / len(model_names)
        print(f"  • Average Model Name Length: {avg_length:.1f} characters")
    
    print("=" * 50)

def display_recent_additions(headers, data, limit=5):
    """Display recent additions to the collection"""
    print(f"\n🆕 RECENT ADDITIONS (Last {limit}) 🆕")
    print("=" * 50)
    
    if len(data) == 0:
        print("📭 No cars in collection yet!")
        return
    
    # Show the last N entries
    recent_cars = data[-limit:] if len(data) >= limit else data
    
    for i, car in enumerate(recent_cars, 1):
        serial_no = car[0] if len(car) > 0 else "N/A"
        model_name = car[1] if len(car) > 1 else "N/A"
        series = car[2] if len(car) > 2 else "N/A"
        print(f"  {i}. S.No: {serial_no} | {model_name} | {series}")
    
    print("=" * 50)

def display_collection_goals(headers, data):
    """Display collection goals and progress"""
    print(f"\n🎯 COLLECTION GOALS 🎯")
    print("=" * 50)
    
    total_cars = len(data)
    
    # Milestone goals
    milestones = [10, 25, 50, 100, 250, 500, 1000]
    next_milestone = None
    
    for milestone in milestones:
        if total_cars < milestone:
            next_milestone = milestone
            break
    
    if next_milestone:
        remaining = next_milestone - total_cars
        print(f"🎯 Next Milestone: {next_milestone} cars")
        print(f"📈 Cars needed: {remaining}")
        print(f"📊 Progress: {(total_cars/next_milestone)*100:.1f}%")
    else:
        print(f"🏆 Congratulations! You've exceeded all standard milestones!")
        print(f"🚀 You have {total_cars} cars - that's an amazing collection!")
    
    print("=" * 50)

def create_simple_chart(series_counts):
    """Create a simple text-based chart"""
    if not series_counts:
        return
    
    print(f"\n📊 SERIES DISTRIBUTION CHART 📊")
    print("=" * 50)
    
    max_count = max(series_counts.values())
    max_bar_length = 30
    
    for series, count in series_counts.most_common():
        bar_length = int((count / max_count) * max_bar_length)
        bar = "█" * bar_length
        print(f"{series:<25} │{bar} {count}")
    
    print("=" * 50)

def main():
    """Main statistics function"""
    print("🔍 HOT WHEELS COLLECTION STATISTICS 🔍")
    print("=" * 60)
    
    # Load data
    headers, data = load_data()
    if headers is None:
        return
    
    # Display different types of statistics
    display_basic_stats(headers, data)
    display_detailed_stats(headers, data)
    display_recent_additions(headers, data)
    display_collection_goals(headers, data)
    
    # Create visual chart if we have series data
    if len(headers) >= 3 and data:
        series_column = 2
        series_counts = Counter(row[series_column] for row in data if row[series_column])
        create_simple_chart(series_counts)
    
    print(f"\n✅ Statistics generated at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

if __name__ == "__main__":
    main()
